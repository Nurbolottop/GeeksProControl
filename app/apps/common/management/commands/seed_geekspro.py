"""Импорт реальных данных GeeksPro из Google Sheets (файл app/data/geekspro_sheets.xlsx).

Полностью очищает базу (кроме суперпользователей) и загружает:
- проекты потоков 10–13 со статусами, этапами и ссылками (ручная разметка
  по цветам таблицы — цвета в xlsx-значениях не сохраняются);
- стажёров из «Ведомость 12/13» с направлением, статусом, филиалом и проектом;
- штат/тимлидов из «Справочник» и листа «тимлиды»;
- учебные группы из листа « план ».

    python manage.py seed_geekspro
"""
import datetime
import re

import openpyxl
from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.audit.models import AuditLog
from apps.clients.models import Client, ClientContact
from apps.documents import services as doc_services
from apps.documents.models import CONTRACT, Document, DocumentType, REQUIREMENTS
from apps.flows.models import Flow
from apps.interns.models import Intern, InternEvaluation, InternStatus
from apps.meetings.models import Meeting
from apps.notifications.models import Notification
from apps.projects.models import (
    Project,
    ProjectStage,
    ProjectStageKey,
    ProjectStatus,
    ProjectStatusHistory,
    ProjectType,
    lifecycle_stages,
)
from apps.reports.models import KPISnapshot, WeeklyReport
from apps.resources.models import PlannedProject
from apps.risks.models import Risk
from apps.tasks.models import Task
from apps.teams.models import TeamMember, TeamRole
from apps.training.models import Specialization, TrainingGroup

User = get_user_model()
D = datetime.date

XLSX_PATH = settings.BASE_DIR / 'data' / 'geekspro_sheets.xlsx'

WEB = 'Web-сайт'
MOBILE = 'Mobile App'

SPEC_NAMES = ['Backend', 'Frontend', 'UX/UI', 'Mobile', 'Testing/QA', 'PM', 'DevOps']
SPEC_MAP = {
    'backend': 'Backend', 'бэкенд': 'Backend',
    'frontend': 'Frontend', 'фронтенд': 'Frontend',
    'ux/ui': 'UX/UI', 'mobile': 'Mobile', 'flatter': 'Mobile',
    'flutter': 'Mobile', 'testing': 'Testing/QA', 'qа': 'Testing/QA',
    'qa': 'Testing/QA', 'pm': 'PM', 'project manager': 'PM',
    'devops-инженер': 'DevOps', 'devops': 'DevOps',
}

ROLE_BY_SPEC = {
    'Backend': TeamRole.BACKEND, 'Frontend': TeamRole.FRONTEND,
    'UX/UI': TeamRole.UXUI, 'Mobile': TeamRole.MOBILE,
    'Testing/QA': TeamRole.QA, 'PM': TeamRole.PROJECT_MANAGER,
    'DevOps': TeamRole.OTHER,
}

# Короткие имена из листов → полные ФИО из «Справочника»/ведомостей
ALIASES = {
    'Эрбол': 'Курманбеков Эрбол',
    'Саидахмат': 'Талантбек уулу Саидахмад',
    'Баэль': 'Мийзамбеков Баэль',
    'Байэл': 'Мийзамбеков Баэль',
    'Жамалодин': 'Сабиржанов Жамолдин',
    'Нурсултан': 'Саимбетов Нурсултан',
    'Жумабек': 'Салиев Жумабек',
    'Билол': 'Джоробаев Билолдин',
    'Ислам': 'Нуралиев Ислам',
    'Бексултан': 'Тагаев Бексултан',
    'Умар': 'Алапаев Умар',
    'Айназик': 'Калдарова Айназик',
    'Имран': 'Насиза Имран',
    'Адэль': 'Бакытбекова Адэль',
    'Мардон': 'Иминжан уулу Мардон',
    'Кудбухон': 'Исакова Кудбухон',
}

# Штат/тимлиды, которых нет в ведомостях: имя → (направление, город)
STAFF = {
    'Саимбетов Нурсултан': ('DevOps', 'Ош'),
    'Курманбеков Эрбол': ('Backend', 'Ош'),
    'Талантбек уулу Саидахмад': ('Backend', 'Ош'),
    'Мийзамбеков Баэль': ('Frontend', 'Ош'),
    'Сабиржанов Жамолдин': ('Mobile', 'Ош'),
    'Салиев Жумабек': ('UX/UI', 'Ош'),
    'Джоробаев Билолдин': ('UX/UI', 'Бишкек'),
    'Исакова Кудбухон': ('UX/UI', 'Ош'),
    'Нуралиев Ислам': ('Frontend', 'Бишкек'),
    'Абдикалилов Ислам': ('Frontend', 'Бишкек'),
    'Иминжан уулу Мардон': ('Backend', 'Бишкек'),
    'Эмиль': ('Testing/QA', 'Бишкек'),
    'Азамат': ('UX/UI', 'Бишкек'),
    'Бекназар': ('Backend', 'Бишкек'),
    'Адина': ('Frontend', 'Бишкек'),
    'Адинай': ('Frontend', 'Бишкек'),
    'Кубанычбек': ('Frontend', 'Бишкек'),
    'Абдукарим': ('Frontend', 'Ош'),
    'Анабел': ('PM', 'Бишкек'),
    'Айкен': ('PM', 'Бишкек'),
    'Кубаныч': ('PM', 'Бишкек'),
    'Эркинай': ('PM', 'Бишкек'),
    'Салия': ('PM', 'Бишкек'),
    'Салтанат': ('PM', 'Бишкек'),
    'Айзирек': ('PM', 'Бишкек'),
    'Перизат': ('PM', 'Бишкек'),
    'Бурулай': ('PM', 'Бишкек'),
    'Арууке': ('PM', 'Бишкек'),
    'Чолпон': ('PM', 'Бишкек'),
    'Айгерим': ('PM', 'Бишкек'),
    'Адилет': ('PM', 'Бишкек'),
    'Аяна': ('PM', 'Бишкек'),
}

# PM по проектам (лист «Ресурсы проектов»)
PROJECT_PM = {
    'technowomen': 'Салтанат', 'Кыргыз экспорт': 'Айзирек',
    'happybaby': 'Перизат', 'Абир Финанс': 'Бурулай', 'USTAT': 'Арууке',
    'MedScan': 'Айкен', 'OkuuKitebi': 'Кубаныч', 'BeStyle': 'Эркинай',
    'NBKR': 'Кубаныч', 'EkiAl': 'Кубаныч',
    'Гикс Про КЖ': 'Адэль', 'ЛогоПарк': 'Чолпон', 'ПроМонтаж': 'Айгерим',
    'Энактас КЖ': 'Адилет', 'БиКлин': 'Кубаныч', 'Умай': 'Аяна',
}

# Дополнительные участники (лист «Ресурсы проектов», потоки 10–11)
EXTRA_TEAM = {
    'MedScan': [('Азамат', 'uxui'), ('Умар', 'backend'), ('Бекназар', 'backend'),
                ('Адина', 'frontend'), ('Баэль', 'frontend')],
    'OkuuKitebi': [('Азамат', 'uxui'), ('Умар', 'backend'), ('Адина', 'frontend')],
    'BeStyle': [('Азамат', 'uxui'), ('Умар', 'backend'), ('Адина', 'frontend')],
    'NBKR': [('Азамат', 'uxui'), ('Умар', 'backend'), ('Адина', 'frontend'),
             ('Эмиль', 'qa')],
    'EkiAl': [('Азамат', 'uxui'), ('Умар', 'backend'), ('Адина', 'frontend'),
              ('Баэль', 'mobile'), ('Имран', 'qa'), ('Жамалодин', 'mobile')],
    'technowomen': [('Айназик', 'frontend'), ('Нурсултан', 'other')],
    'Кыргыз экспорт': [('Айназик', 'frontend'), ('Нурсултан', 'other')],
    'happybaby': [('Адинай', 'frontend'), ('Айназик', 'frontend'),
                  ('Нурсултан', 'other')],
    'Абир Финанс': [('Адинай', 'frontend'), ('Айназик', 'frontend'),
                    ('Нурсултан', 'other')],
    'USTAT': [('Адинай', 'frontend'), ('Айназик', 'frontend')],
    'БАТ-МУ': [('Айназик', 'frontend'), ('Нурсултан', 'other')],
    'Гикс Про КЖ': [('Баэль', 'frontend')],
    'ЛогоПарк': [('Баэль', 'frontend')],
    'ПроМонтаж': [('Баэль', 'frontend')],
    'Энактас КЖ': [('Баэль', 'frontend')],
    'Умай': [('Алишер', 'backend'), ('Нурсултан', 'other')],
    'КИТ форум': [('Нурсултан', 'other')],
}

# Проекты (лист «проекты» + «Ресурсы проектов»; статусы/этапы — по цветам таблицы)
PROJECTS = [
    # -------- Поток 13 — в работе --------
    dict(name='Олимпийская школа', city='Бишкек', type=WEB, flow=13,
         contract=D(2026, 5, 18), deadline=D(2026, 8, 16),
         status='active', stage=ProjectStageKey.BACKEND, progress=60),
    dict(name='Омур клиника', city='Ош', type=WEB, flow=13,
         contract=D(2026, 4, 14), deadline=D(2026, 7, 13),
         status='active', stage=ProjectStageKey.TESTING, progress=75),
    dict(name='ОБА', city='Ош', type=WEB, flow=13,
         contract=D(2026, 3, 1), deadline=D(2026, 7, 3),
         status='active', stage=ProjectStageKey.BACKEND, progress=65),
    dict(name='Ислам-медресе (Абдушукур Нарматов)', city='Ош', type=WEB, flow=13,
         contract=D(2026, 5, 19), deadline=D(2026, 8, 17),
         status='active', stage=ProjectStageKey.BACKEND, progress=60),
    dict(name='Мин просвещения', city='Бишкек', type=WEB, flow=13,
         contract=D(2026, 6, 23), deadline=D(2026, 9, 21),
         status='active', stage=ProjectStageKey.BACKEND, progress=45),
    dict(name='Учкун', city='Бишкек', type=WEB, flow=13,
         status='active', stage=ProjectStageKey.DESIGN, progress=25),
    dict(name='Полароид', city='Бишкек', type=WEB, flow=13,
         status='active', stage=ProjectStageKey.REQUIREMENTS, progress=5),

    # -------- Поток 12 --------
    dict(name='Гикс Про КЖ', city='Бишкек', type=WEB, flow=12,
         contract=D(2026, 3, 9), deadline=D(2026, 5, 31), done=D(2026, 5, 31),
         status='completed', stage=ProjectStageKey.COMPLETED, progress=100,
         has_contract=True, has_tz=True),
    dict(name='ЛогоПарк', city='Бишкек', type=WEB, flow=12,
         contract=D(2026, 3, 9),
         status='cancelled', stage=ProjectStageKey.REQUIREMENTS, progress=10,
         has_contract=True, has_tz=True),
    dict(name='ПроМонтаж', city='Бишкек', type=WEB, flow=12,
         contract=D(2026, 3, 9), deadline=D(2026, 5, 30), done=D(2026, 5, 30),
         status='completed', stage=ProjectStageKey.COMPLETED, progress=100,
         has_contract=True, has_tz=True),
    dict(name='Энактас КЖ', city='Бишкек', type=WEB, flow=12,
         contract=D(2026, 3, 9), deadline=D(2026, 5, 6), done=D(2026, 5, 15),
         status='completed', stage=ProjectStageKey.COMPLETED, progress=100,
         has_contract=True, has_tz=True),
    dict(name='БиКлин', city='Бишкек', type=MOBILE, flow=12,
         deadline=D(2026, 7, 31),
         status='active', stage=ProjectStageKey.TESTING, progress=80),

    # -------- Поток 11 --------
    dict(name='MedScan', city='Бишкек', type=WEB, flow=11,
         contract=D(2026, 2, 19), deadline=D(2026, 6, 19),
         status='active', stage=ProjectStageKey.REWORK, progress=90,
         staging='https://medscan.geekspro.kg',
         gitlab='https://gitlab.geeks.kg/medscan', has_tz=True,
         note='Сроки переносились: 31.03 → 05.04. Продакшена нет.'),
    dict(name='OkuuKitebi', city='Бишкек', type=WEB, flow=11,
         contract=D(2025, 11, 1), deadline=D(2026, 4, 15),
         status='active', stage=ProjectStageKey.DELIVERY, progress=95,
         staging='https://okuukitebi.geekspro.kg',
         production='https://okuukitebi.edu.kg',
         gitlab='https://gitlab.geeks.kg/okuukitebi', has_tz=True),
    dict(name='BeStyle', city='Бишкек', type=WEB, flow=11,
         contract=D(2026, 1, 5), deadline=D(2026, 4, 30),
         status='active', stage=ProjectStageKey.PRODUCTION, progress=90,
         staging='https://bestyle.geekspro.kg',
         gitlab='https://gitlab.geeks.kg/bestyle', has_tz=True),
    dict(name='NBKR', city='Бишкек', type=WEB, flow=11,
         contract=D(2026, 2, 21), deadline=D(2026, 3, 30), done=D(2026, 3, 30),
         status='completed', stage=ProjectStageKey.COMPLETED, progress=100,
         staging='https://nbkr.test.com.kg',
         gitlab='https://gitlab.geeks.kg/nbkr',
         note='Вышел в прод 30.03. ТЗ не было.'),
    dict(name='EkiAl', city='Бишкек', type=MOBILE, flow=11,
         contract=D(2025, 11, 21), deadline=D(2026, 4, 30),
         status='refused', stage=ProjectStageKey.BACKEND, progress=40,
         gitlab='https://gitlab.geeks.kg/ekial', has_tz=True),

    # -------- Поток 10 --------
    dict(name='technowomen', city='Бишкек', type=WEB, flow=10,
         contract=D(2025, 8, 8), deadline=D(2026, 5, 31),
         status='active', stage=ProjectStageKey.PRODUCTION, progress=85,
         staging='https://technowoman.test.com.kg',
         gitlab='https://gitlab.geeks.kg/technowoman', has_tz=True,
         note='Договора нет. Доступа к продакшену нет.'),
    dict(name='Кыргыз экспорт', city='Бишкек', type=WEB, flow=10,
         contract=D(2025, 8, 15), deadline=D(2026, 7, 1),
         status='active', stage=ProjectStageKey.PRODUCTION, progress=85,
         staging='https://kyrgyz.geekspro.kg',
         gitlab='https://gitlab.geeks.kg/kyrgyzexport',
         note='Договора нет. Доступ к продакшену получили.'),
    dict(name='happybaby', city='Бишкек', type=WEB, flow=10,
         contract=D(2025, 8, 8), deadline=D(2026, 7, 20),
         status='active', stage=ProjectStageKey.PRODUCTION, progress=85,
         staging='https://happybaby.geekspro.kg',
         gitlab='https://gitlab.geeks.kg/happy-baby', has_tz=True,
         note='Доступа к продакшену нет — студия должна дать.'),
    dict(name='Абир Финанс', city='Бишкек', type=WEB, flow=10,
         contract=D(2025, 8, 4), deadline=D(2026, 7, 20),
         status='active', stage=ProjectStageKey.PRODUCTION, progress=85,
         staging='https://abiyir.geekspro.kg',
         gitlab='https://gitlab.geeks.kg/abiyirfinance', has_tz=True,
         note='Сервер клиента не подходит.'),
    dict(name='USTAT', city='Бишкек', type=WEB, flow=10,
         status='paused', stage=ProjectStageKey.BACKEND, progress=50,
         has_tz=True, note='Сроки под вопросом, договора нет.'),
    dict(name='БАТ-МУ', city='Бишкек', type=WEB, flow=10,
         contract=D(2025, 8, 8), deadline=D(2026, 1, 21),
         status='refused', stage=ProjectStageKey.PRODUCTION, progress=90,
         staging='https://bsu.geekspro.kg',
         gitlab='https://gitlab.geeks.kg/batken-state-university', has_tz=True,
         note='Не вышел в прод — проект остался на стейдже.'),
    dict(name='Кок-Бел', city='Ош', type=WEB, flow=10,
         contract=D(2025, 10, 1), deadline=D(2026, 1, 31),
         status='active', stage=ProjectStageKey.REWORK, progress=80),
    dict(name='АДДК-Колледж', city='Ош', type=WEB, flow=10, client='АДДК',
         contract=D(2025, 12, 1), deadline=D(2026, 2, 20), done=D(2026, 2, 20),
         status='completed', stage=ProjectStageKey.COMPLETED, progress=100),
    dict(name='АДДК-Академия', city='Ош', type=WEB, flow=10, client='АДДК',
         contract=D(2025, 12, 1), deadline=D(2026, 2, 20), done=D(2026, 2, 20),
         status='completed', stage=ProjectStageKey.COMPLETED, progress=100),
    dict(name='АДДК-Садик', city='Ош', type=WEB, flow=10, client='АДДК',
         contract=D(2026, 4, 1), deadline=D(2026, 6, 30),
         status='active', stage=ProjectStageKey.BACKEND, progress=70),
    dict(name='АДДК-Мектеп', city='Ош', type=WEB, flow=10, client='АДДК',
         contract=D(2026, 4, 1), deadline=D(2026, 6, 30),
         status='active', stage=ProjectStageKey.BACKEND, progress=70),

    # -------- Вне потоков --------
    dict(name='Умай', city='Бишкек', type=WEB, flow=13,
         status='active', stage=ProjectStageKey.BACKEND, progress=40,
         gitlab='https://gitlab.geeks.kg/umai', has_tz=True),
    dict(name='КИТ форум', city='Бишкек', type=WEB, flow=13,
         status='active', stage=ProjectStageKey.NEW, progress=0),
]

# Названия проектов в листах → канонические названия
PROJECT_NAME_ALIASES = {
    'ислам - медресе(а.нарматов)': 'Ислам-медресе (Абдушукур Нарматов)',
    'ислам - медресе (абдүшүкүр нарматов)': 'Ислам-медресе (Абдушукур Нарматов)',
    'палароид': 'Полароид',
    'аддк-коледж': 'АДДК-Колледж', 'аддк-академия': 'АДДК-Академия',
    'аддк - садик': 'АДДК-Садик', 'аддк - мектеп': 'АДДК-Мектеп',
    'geekspro': 'Гикс Про КЖ', 'logopark': 'ЛогоПарк',
    'promontage': 'ПроМонтаж', 'enactus': 'Энактас КЖ', 'biclean': 'БиКлин',
}

STATUS_MAP = {
    'active': ProjectStatus.ACTIVE, 'paused': ProjectStatus.PAUSED,
    'completed': ProjectStatus.COMPLETED, 'cancelled': ProjectStatus.CANCELLED,
    'refused': ProjectStatus.REFUSED,
}


def normalize(value) -> str:
    return re.sub(r'\s+', ' ', str(value or '')).strip()


class Command(BaseCommand):
    help = 'Очищает базу и импортирует реальные данные GeeksPro из xlsx'

    @transaction.atomic
    def handle(self, *args, **options):
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        self._wipe()
        self.specs = {
            name: Specialization.objects.get_or_create(name=name)[0]
            for name in SPEC_NAMES
        }
        self.types = {
            name: ProjectType.objects.get_or_create(name=name)[0]
            for name in (WEB, MOBILE)
        }
        self.interns: dict[str, Intern] = {}
        self.memberships: set[tuple[int, int]] = set()

        self.projects = self._create_projects()
        self._create_staff()
        roster_count = self._import_roster(
            wb['Ведомость 13'], name_col=5, roster_flow=13,
        )
        roster_count += self._import_roster(
            wb['Ведомость 12'], name_col=6, roster_flow=12,
        )
        self._assign_team_leads(wb['тимлиды'])
        self._assign_pms_and_extras()
        groups = self._import_training_groups(wb[' план '])
        self._distribute_workloads()
        # Последовательные внутренние ID
        for index, project in enumerate(Project.objects.order_by('pk'), 1):
            Project.objects.filter(pk=project.pk).update(code=f'GP-{index:04d}')

        self.stdout.write(self.style.SUCCESS(
            f'Импорт завершён: {len(self.projects)} проектов, '
            f'{Client.objects.count()} клиентов, '
            f'{Intern.objects.count()} человек ({roster_count} из ведомостей), '
            f'{TeamMember.objects.count()} назначений, '
            f'{groups} учебных групп.',
        ))

    # ------------------------------------------------------------------
    def _wipe(self):
        self.stdout.write('Очистка базы…')
        for model in (
            AuditLog, Notification, Risk, WeeklyReport, KPISnapshot,
            Meeting, Document, Task, TeamMember, InternEvaluation, Intern,
            TrainingGroup, PlannedProject, ProjectStatusHistory, ProjectStage,
            Project, ClientContact, Client, Flow,
        ):
            model.objects.all().delete()
        User.objects.filter(is_superuser=False).delete()

    # ------------------------------------------------------------------
    def _create_projects(self) -> dict[str, Project]:
        doc_services.ensure_default_types()
        contract_type = DocumentType.objects.get(code=CONTRACT)
        tz_type = DocumentType.objects.get(code=REQUIREMENTS)
        clients: dict[str, Client] = {}
        projects: dict[str, Project] = {}

        # Потоки: 13 и новее — активные, остальные завершены
        flows = {
            number: Flow.objects.create(
                number=number,
                status=Flow.Status.ACTIVE if number >= 13 else Flow.Status.FINISHED,
            )
            for number in sorted({item['flow'] for item in PROJECTS})
        }
        self.flows = flows
        flow_counters = {number: 0 for number in flows}

        for data in PROJECTS:
            client_name = data.get('client', data['name'])
            if client_name not in clients:
                clients[client_name] = Client.objects.create(
                    organization=client_name, city=data['city'],
                )
            project = Project.objects.create(
                name=data['name'],
                client=clients[client_name],
                city=data['city'],
                project_type=self.types[data['type']],
                contract_date=data.get('contract'),
                start_date=data.get('contract'),
                planned_end_date=data.get('deadline'),
                actual_end_date=data.get('done'),
                status=STATUS_MAP[data['status']],
                current_stage=data['stage'],
                progress=data['progress'],
                staging_url=data.get('staging', ''),
                production_url=data.get('production', ''),
                github_url=data.get('gitlab', ''),
                flow=flows[data['flow']],
                number_in_flow=self._next_number(flow_counters, data['flow']),
                head_comment=data.get('note', ''),
            )
            self._create_stages(project, data['stage'])
            ProjectStatusHistory.objects.create(
                project=project, field='created',
                new_value='Импортировано из Google Sheets',
            )
            if data.get('has_contract'):
                Document.objects.create(
                    project=project, doc_type=contract_type, status='signed',
                    is_signed=True, signed_date=data.get('contract'),
                    document_date=data.get('contract'),
                )
            if data.get('has_tz'):
                Document.objects.create(
                    project=project, doc_type=tz_type, status='signed',
                    is_signed=True,
                    signed_date=data.get('contract') or data.get('deadline'),
                    document_date=data.get('contract'),
                )
            projects[data['name']] = project
        return projects

    @staticmethod
    def _next_number(counters: dict, flow_number: int) -> int:
        counters[flow_number] += 1
        return counters[flow_number]

    def _create_stages(self, project: Project, current_stage: str):
        stage_order = lifecycle_stages(project.project_type)
        current_index = (
            stage_order.index(current_stage) if current_stage in stage_order else 0
        )
        stages = []
        for index, key in enumerate(stage_order):
            if project.status == ProjectStatus.COMPLETED or index < current_index:
                stage_status, stage_progress = ProjectStage.Status.DONE, 100
            elif index == current_index:
                stage_status, stage_progress = (
                    ProjectStage.Status.IN_PROGRESS, project.progress,
                )
            else:
                stage_status, stage_progress = ProjectStage.Status.NOT_STARTED, 0
            stages.append(ProjectStage(
                project=project, key=key, order=index,
                status=stage_status, progress=stage_progress,
                start_date=project.start_date if index <= current_index else None,
                end_date=project.actual_end_date
                         if project.status == ProjectStatus.COMPLETED else None,
            ))
        ProjectStage.objects.bulk_create(stages)

    # ------------------------------------------------------------------
    def _get_intern(self, raw_name: str, spec_name: str = 'Backend',
                    city: str = 'Бишкек') -> Intern:
        name = ALIASES.get(normalize(raw_name), normalize(raw_name))
        if name not in self.interns:
            self.interns[name] = Intern.objects.create(
                full_name=name,
                specialization=self.specs.get(spec_name),
                status=InternStatus.ACTIVE,
                city=city,
            )
        return self.interns[name]

    def _create_staff(self):
        for name, (spec, city) in STAFF.items():
            self._get_intern(name, spec, city)
        # Алишер и Бексултан — тимлиды из листа «тимлиды»
        self._get_intern('Алишер', 'Backend', 'Бишкек')
        self._get_intern('Бексултан', 'Backend', 'Бишкек')

    # ------------------------------------------------------------------
    def _resolve_project(self, raw_name) -> Project | None:
        name = normalize(raw_name)
        if not name or name.lower() in ('wfk',):
            return None
        canonical = PROJECT_NAME_ALIASES.get(name.lower(), name)
        for project_name, project in self.projects.items():
            if project_name.lower() == canonical.lower():
                return project
        # частичное совпадение («Ислам - медресе(А.Нарматов)» и т.п.)
        for project_name, project in self.projects.items():
            first = canonical.lower().split(' ')[0].split('-')[0]
            if first and project_name.lower().startswith(first):
                return project
        return None

    def _import_roster(self, ws, *, name_col: int, roster_flow: int | None = None) -> int:
        """Ведомость: A поток, B старт, D/E направление, E/F проект, F/G ФИО…"""
        count = 0
        offset = name_col - 5  # 0 для «Ведомость 13», 1 для «Ведомость 12»
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_name = row[name_col] if len(row) > name_col else None
            if not raw_name or not normalize(raw_name):
                continue
            name = normalize(raw_name)
            spec_raw = normalize(row[3 + offset]).lower()
            spec_name = SPEC_MAP.get(spec_raw, 'Backend')
            status_raw = normalize(row[name_col + 1])
            branch_raw = normalize(row[name_col + 2])
            start = row[1] if isinstance(row[1], datetime.datetime) else None

            if 'Прекратил' in status_raw:
                status = InternStatus.DROPPED
            elif 'Активный' in status_raw:
                status = InternStatus.ACTIVE
            else:
                status = InternStatus.WAITING

            city = 'Ош' if 'Ош' in branch_raw else 'Бишкек'
            if name in self.interns:
                intern = self.interns[name]
            else:
                intern = Intern.objects.create(
                    full_name=name,
                    specialization=self.specs[spec_name],
                    status=status,
                    city=city,
                    branch=branch_raw or '',
                    internship_start_date=start.date() if start else None,
                    flow=self.flows.get(roster_flow),
                )
                self.interns[name] = intern
                count += 1

            project = self._resolve_project(row[4 + offset])
            if project and status != InternStatus.DROPPED:
                self._add_member(
                    project, intern, ROLE_BY_SPEC.get(spec_name, TeamRole.OTHER),
                )
        return count

    # ------------------------------------------------------------------
    def _add_member(self, project: Project, intern: Intern, role: str):
        key = (project.pk, intern.pk)
        if key in self.memberships:
            return
        self.memberships.add(key)
        is_active_project = project.status in (
            ProjectStatus.ACTIVE, ProjectStatus.PAUSED,
        )
        TeamMember.objects.create(
            project=project, intern=intern, role=role,
            workload=0,  # проставим при распределении
            joined_at=project.start_date,
            left_at=None if is_active_project else (
                project.actual_end_date or project.planned_end_date
            ),
            status=TeamMember.Status.ACTIVE if is_active_project
                   else TeamMember.Status.LEFT,
        )

    def _assign_team_leads(self, ws):
        """Лист «тимлиды»: направленческие тимлиды по каждому проекту."""
        for row in ws.iter_rows(min_row=2, values_only=True):
            project = self._resolve_project(row[0])
            if project is None:
                continue
            for col, spec in ((2, 'PM'), (3, 'UX/UI'), (4, 'Backend'),
                              (5, 'Frontend'), (6, 'Mobile'),
                              (7, 'Testing/QA'), (8, 'Backend')):
                raw = normalize(row[col]) if len(row) > col else ''
                if not raw:
                    continue
                for person in re.split(r'[/,]', raw):
                    person = normalize(person)
                    if person:
                        intern = self._get_intern(person, spec)
                        self._add_member(project, intern, TeamRole.TEAM_LEAD)

    def _assign_pms_and_extras(self):
        for project_name, pm_name in PROJECT_PM.items():
            project = self.projects.get(project_name)
            if project:
                intern = self._get_intern(pm_name, 'PM')
                self._add_member(project, intern, TeamRole.PROJECT_MANAGER)
        for project_name, members in EXTRA_TEAM.items():
            project = self.projects.get(project_name)
            if not project:
                continue
            role_map = {
                'pm': TeamRole.PROJECT_MANAGER, 'backend': TeamRole.BACKEND,
                'frontend': TeamRole.FRONTEND, 'uxui': TeamRole.UXUI,
                'mobile': TeamRole.MOBILE, 'qa': TeamRole.QA,
                'other': TeamRole.OTHER,
            }
            for person, role in members:
                intern = self._get_intern(person)
                self._add_member(project, intern, role_map[role])

    # ------------------------------------------------------------------
    def _import_training_groups(self, ws) -> int:
        """Лист « план »: направление, филиал, № группы, дата окончания…"""
        created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            end = row[4] if len(row) > 4 else None
            if not isinstance(end, datetime.datetime):
                continue
            spec_raw = normalize(row[1]).lower()
            spec_name = SPEC_MAP.get(spec_raw)
            if not spec_name:
                continue
            branch = normalize(row[2])
            number = normalize(row[3]) or f'{spec_name} ({branch})'
            students = int(row[5]) if isinstance(row[5], (int, float)) else 0
            expected = int(row[6]) if isinstance(row[6], (int, float)) else 0
            TrainingGroup.objects.create(
                number=number,
                specialization=self.specs[spec_name],
                branch=branch,
                end_date=end.date(),
                students_count=students,
                expected_interns=expected,
            )
            created += 1
        return created

    # ------------------------------------------------------------------
    def _distribute_workloads(self):
        """Загрузка: 100% делится между активными проектами человека."""
        active = TeamMember.objects.filter(
            status=TeamMember.Status.ACTIVE, intern__isnull=False,
        )
        counts: dict[int, int] = {}
        for member in active:
            counts[member.intern_id] = counts.get(member.intern_id, 0) + 1
        for member in active:
            n = counts[member.intern_id]
            member.workload = 80 if n == 1 else max(15, min(50, 100 // n))
            member.save(update_fields=['workload'])
