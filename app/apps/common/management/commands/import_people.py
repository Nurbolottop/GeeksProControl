"""Импорт стажёров, тимлидов и PM для проектов первого потока.

Источник — рабочая таблица app/data/geekspro_sheets_v2.xlsx
(листы «Ведомость 13», «Ведомость 12», «тимлиды», «Ресурсы проектов»).

    python manage.py import_people
"""
import re

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.flows.models import Group
from apps.interns.models import Intern, InternStatus
from apps.projects.models import Project
from apps.teams.models import TeamMember, TeamRole
from apps.training.models import Specialization

XLSX = settings.BASE_DIR / 'data' / 'geekspro_sheets_v2.xlsx'

# Как проект называется в таблице → как он называется у нас
PROJECT_ALIASES = {
    'омур клиника': 'Омур',
    'омур': 'Омур',
    'оба': 'ОБА',
    'балажан': 'Балажан',
    'билим ордо': 'БилимОрдо',
    'билимордо': 'БилимОрдо',
    'умай': 'Umai',
    'umai': 'Umai',
    'биклин': 'Биклин',
    'biclean': 'Биклин',
    'олимпийская школа': 'Олимпийская школа',
    'ислам - медресе(а.нарматов)': 'Медресе',
    'ислам - медресе (абдүшүкүр нарматов)': 'Медресе',
    'медресе': 'Медресе',
    'bestyle': 'Вистайл',
    'вистайл': 'Вистайл',
    'учкун': 'Учкун',
    'мин просвещения': 'Агартуу',
    'агартуу': 'Агартуу',
    'wfk': 'ВФК',
    'вфк': 'ВФК',
    'энактас кж': 'Энактус',
    'энактус': 'Энактус',
    'enactus': 'Энактус',
    'промонтаж': 'ПроМонтаж',
    'promontage': 'ПроМонтаж',
    'туризм': 'Туризм',
}

SPEC_MAP = {
    'backend': 'Backend', 'бэкенд': 'Backend', 'бекенд': 'Backend',
    'frontend': 'Frontend', 'фронтенд': 'Frontend',
    'ux/ui': 'UX/UI', 'дизайн': 'UX/UI',
    'mobile': 'Mobile', 'flutter': 'Mobile', 'flatter': 'Mobile',
    'testing': 'Testing/QA', 'qa': 'Testing/QA', 'тестирование': 'Testing/QA',
    'pm': 'PM', 'project manager': 'PM',
    'devops': 'DevOps', 'devops-инженер': 'DevOps',
}

ROLE_BY_SPEC = {
    'Backend': TeamRole.BACKEND, 'Frontend': TeamRole.FRONTEND,
    'UX/UI': TeamRole.UXUI, 'Mobile': TeamRole.MOBILE,
    'Testing/QA': TeamRole.QA, 'PM': TeamRole.PROJECT_MANAGER,
    'DevOps': TeamRole.OTHER,
}


def clean(value) -> str:
    return re.sub(r'\s+', ' ', str(value or '')).strip()


class Command(BaseCommand):
    help = 'Импортирует людей проектов первого потока из таблицы'

    @transaction.atomic
    def handle(self, *args, **options):
        wb = openpyxl.load_workbook(XLSX, data_only=True)
        self.specs = {
            name: Specialization.objects.get_or_create(name=name)[0]
            for name in ('Backend', 'Frontend', 'UX/UI', 'Mobile',
                         'Testing/QA', 'PM', 'DevOps')
        }
        self.groups = {
            group.project.name: group
            for group in Group.objects.select_related('project')
            if group.project_id
        }
        self.people: dict[str, Intern] = {}
        self.links: set[tuple[int, int]] = set()

        interns = self._import_roster(wb['Ведомость 13'], name_col=5)
        interns += self._import_roster(wb['Ведомость 12'], name_col=6)
        leads = self._import_leads(wb['тимлиды'])
        pms = self._import_pms(wb['Ресурсы проектов'])

        self.stdout.write(self.style.SUCCESS(
            f'Загружено: {Intern.objects.count()} человек '
            f'({interns} из ведомостей, {leads} тимлидов, {pms} PM), '
            f'{TeamMember.objects.count()} назначений в командах.',
        ))

    # ------------------------------------------------------------------
    def _project_group(self, raw_name) -> Group | None:
        name = clean(raw_name).lower()
        if not name:
            return None
        our_name = PROJECT_ALIASES.get(name)
        if our_name is None:
            for alias, target in PROJECT_ALIASES.items():
                if alias in name or name in alias:
                    our_name = target
                    break
        return self.groups.get(our_name) if our_name else None

    def _person(self, full_name: str, spec_name: str, city: str = 'Бишкек',
                branch: str = '', status: str = InternStatus.ACTIVE,
                phone: str = '', telegram: str = '') -> Intern:
        name = clean(full_name)
        if name in self.people:
            return self.people[name]
        intern, _ = Intern.objects.get_or_create(
            full_name=name,
            defaults={
                'specialization': self.specs.get(spec_name),
                'status': status,
                'city': city,
                'branch': branch,
                'phone': phone,
                'comment': f'Telegram: {telegram}' if telegram else '',
            },
        )
        self.people[name] = intern
        return intern

    def _add_member(self, group: Group, intern: Intern, role: str) -> bool:
        key = (group.pk, intern.pk)
        if key in self.links:
            return False
        self.links.add(key)
        if TeamMember.objects.filter(group=group, intern=intern).exists():
            return False
        TeamMember.objects.create(
            group=group, project=group.project, intern=intern, role=role,
            workload=50, status=TeamMember.Status.ACTIVE,
        )
        return True

    # ------------------------------------------------------------------
    def _import_roster(self, ws, *, name_col: int) -> int:
        """Ведомость: направление, проект, ФИО, статус, филиал."""
        offset = name_col - 5
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_name = row[name_col] if len(row) > name_col else None
            if not clean(raw_name):
                continue
            group = self._project_group(row[4 + offset])
            if group is None:
                continue

            spec_name = SPEC_MAP.get(clean(row[3 + offset]).lower(), 'Backend')
            status_raw = clean(row[name_col + 1])
            branch = clean(row[name_col + 2])
            if 'Прекратил' in status_raw:
                status = InternStatus.DROPPED
            elif 'Активный' in status_raw:
                status = InternStatus.ACTIVE
            else:
                status = InternStatus.WAITING

            # Телефон и телеграм — в свободных колонках после статуса
            phone, telegram = '', ''
            for cell in row[name_col + 3:]:
                text = clean(cell)
                if not text or text.startswith('---'):
                    continue
                if '@' in text and not telegram:
                    telegram = text.split()[0]
                digits = re.sub(r'\D', '', text)
                if len(digits) >= 9 and not phone:
                    phone = digits

            intern = self._person(
                raw_name, spec_name,
                city='Ош' if 'Ош' in branch else 'Бишкек',
                branch=branch, status=status, phone=phone, telegram=telegram,
            )
            if status != InternStatus.DROPPED:
                count += int(self._add_member(
                    group, intern, ROLE_BY_SPEC.get(spec_name, TeamRole.OTHER),
                ))
        return count

    def _import_leads(self, ws) -> int:
        """Лист «тимлиды»: тимлиды направлений по проектам."""
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            group = self._project_group(row[0])
            if group is None:
                continue
            for col, spec in ((2, 'PM'), (3, 'UX/UI'), (4, 'Backend'),
                              (5, 'Frontend'), (6, 'Mobile'),
                              (7, 'Testing/QA'), (8, 'Backend')):
                raw = clean(row[col]) if len(row) > col else ''
                if not raw:
                    continue
                for person in re.split(r'[/,]', raw):
                    person = clean(person)
                    if not person:
                        continue
                    intern = self._person(person, spec)
                    count += int(self._add_member(
                        group, intern, TeamRole.TEAM_LEAD,
                    ))
        return count

    def _import_pms(self, ws) -> int:
        """Лист «Ресурсы проектов»: PM проектов (колонка S)."""
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            group = self._project_group(row[2] if len(row) > 2 else None)
            if group is None:
                continue
            pm_name = clean(row[18]) if len(row) > 18 else ''
            if not pm_name:
                continue
            intern = self._person(pm_name, 'PM')
            count += int(self._add_member(
                group, intern, TeamRole.PROJECT_MANAGER,
            ))
        return count
